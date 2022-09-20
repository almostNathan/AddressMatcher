from distutils.command.build import build
from openpyxl import Workbook, Worksheet, load_workbook
from openpyxl.utils import get_column_letter
import tkinter
from tkinter import filedialog
import os


class MySheet(Worksheet):

    def __init__(self):
        super().__init__()
        self.headerlist = []
        

    def headerList(self):
        for col in range(1, self.active.max_column+1):
            cell = self.active[get_column_letter(col)+"1"].value
            self.headerList.append(cell)
        return self.headerList


workbook = load_workbook("AceTest.xlsx")
worksheet = MySheet(workbook.active)
print(type(worksheet))