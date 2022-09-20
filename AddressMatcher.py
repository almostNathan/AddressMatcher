from distutils.command.build import build
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import tkinter
from tkinter import filedialog
import os


#parameters 
#   sheets = array of sheets being matched
#   addressLists = array of [address, zip]
#   matches = array of matched rows [[a,b],[c,d],[x,y]]
#append matched addresses to rows
def writeAddresses(sheets, matches):
    lastColumns = []
    for sheet in sheets:
        lastColumns.append(get_column_letter(sheet.max_column+1))
        print(lastColumns)
    for item in matches:

        sheets[0][lastColumns[0]+str(item[0]+1)].value = sheets[1][lastColumns[1]+str(item[1]+1)].value
        sheets[1][lastColumns[1]+str(item[1]+1)].value = sheets[0][lastColumns[0]+str(item[0]+1)].value



#takes list if strings and 
#returns list of row matches [[x,y],[a,b],[c,z]]
def getAddresses(sheet, selectedCols):
    maxRows = sheet.max_row
    addressesFull = []
    zipList = []
    for rowNum in range(2,maxRows+1):
        addressesFull.append(sheet[selectedCols[0]+str(rowNum)].value)
        zipList.append(sheet[selectedCols[1]+str(rowNum)].value)
    addressesNum = list(map(lambda x: x.split()[0],addressesFull))

    returnAddressList=[]
    for i in range(0, maxRows-1):
        returnAddressList.append([addressesNum[i],zipList[i]])

    return returnAddressList
    

    
#displays column selector with header names
def buildColumnSelector(headerList):
    window = tkinter.Tk()
    window.title('Select Columns')
    window.geometry('500x500')

    label = tkinter.Label(window, bg='white', width=20, text='Select Address and Zip')
    label.pack()

    #initialize list to hold select options
    colList = []
    #list of tkinter.IntVar()'s
    for i in range(0,len(headerList)):
        colList.append(tkinter.IntVar())

    counter = 0
    for column in headerList:
        tkinter.Checkbutton(window, text=column, variable=colList[counter], onvalue= counter+1, offvalue=0).pack()
        counter+=1
        
    #reference
    #colA = tkinter.Checkbutton(window, text='A', variable=colList[0], onvalue= 1, offvalue=0).pack()

    btn = tkinter.Button(window, text='Submit', command = window.destroy).pack()

    window.mainloop()
    return colList

#takes a list of column numbers and returns list of letters
def getSelectedColLetters(headerList):
    #ask user what columns contain addresses
    selectedCols = buildColumnSelector(headerList)
    #remove not selected columns
    selectedCols = list(filter(lambda x: x.get()!=0,selectedCols))
    #get list of values as column LETTERS
    selectedCols = list(map(lambda x: get_column_letter(x.get()), selectedCols))

    return selectedCols


#get the headers from the sheet
def getHeaders(sheet):
    headerList = []
    for col in range(1, sheet.max_column+1):
        cell = sheet[get_column_letter(col)+"1"].value
        headerList.append(cell)
    return headerList


workbookFilePath = []
#get the workbook filepath from the user, dialog fileselect window
#workbookFilePath.append("C:/Users/natha/Documents/BBB Address Matcher/AceTest.xlsx")
workbookFilePath.append(filedialog.askopenfilename(initialdir=os.getcwd(), title="Select the Excel File", filetypes=(("Excel Files", "*.xlsx"),("All Files", "*.*"))))

#workbookFilePath.append("C:/Users/natha/Documents/BBB Address Matcher/AceTest2.xlsx")
workbookFilePath.append(filedialog.askopenfilename(initialdir=os.getcwd(), title="Select the Excel File", filetypes=(("Excel Files", "*.xlsx"),("All Files", "*.*"))))

workbooks = []
sheets = []
addressZip = []

for index, filePath in enumerate(workbookFilePath):
    workbooks.append(load_workbook(filePath))
    sheets.append(workbooks[index].active)
    headerList = getHeaders(sheets[index])
    selectedCols = getSelectedColLetters(headerList)
    addressZip.append(getAddresses(sheets[index], selectedCols))
    

matches = []
for item1 in addressZip[0]:
    for item2 in addressZip[1]:
        if item1==item2:
           matches.append([addressZip[0].index(item1),addressZip[1].index(item2)]) 


writeAddresses(sheets, matches)

for index, workbook in enumerate(workbooks):
    workbook.save(workbookFilePath[index])
    workbook.close